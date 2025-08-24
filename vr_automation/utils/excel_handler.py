"""
Utilitário para manipulação de arquivos Excel
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Optional, Tuple, Any
import os
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


class ExcelHandler:
    """Classe para manipulação de arquivos Excel"""
    
    def __init__(self):
        """Inicializar handler"""
        self.workbook = None
        self.worksheets = {}
    
    def read_excel_file(self, file_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """
        Ler arquivo Excel e retornar DataFrame
        
        Args:
            file_path: Caminho do arquivo
            sheet_name: Nome da aba (opcional)
            
        Returns:
            DataFrame com os dados
        """
        try:
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)
            
            logger.info(f"Arquivo {file_path} lido com sucesso. Shape: {df.shape}")
            return df
            
        except Exception as e:
            logger.error(f"Erro ao ler arquivo {file_path}: {str(e)}")
            raise
    
    def read_multiple_sheets(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """
        Ler múltiplas abas de um arquivo Excel
        
        Args:
            file_path: Caminho do arquivo
            
        Returns:
            Dicionário com nome da aba e DataFrame
        """
        try:
            excel_file = pd.ExcelFile(file_path)
            sheets = {}
            
            for sheet_name in excel_file.sheet_names:
                sheets[sheet_name] = pd.read_excel(file_path, sheet_name=sheet_name)
            
            logger.info(f"Arquivo {file_path} lido com {len(sheets)} abas")
            return sheets
            
        except Exception as e:
            logger.error(f"Erro ao ler múltiplas abas do arquivo {file_path}: {str(e)}")
            raise
    
    def save_dataframe_to_excel(self, df: pd.DataFrame, file_path: str, 
                               sheet_name: str = "Sheet1", 
                               index: bool = False) -> None:
        """
        Salvar DataFrame em arquivo Excel
        
        Args:
            df: DataFrame para salvar
            file_path: Caminho do arquivo
            sheet_name: Nome da aba
            index: Se deve incluir índice
        """
        try:
            # Criar diretório se não existir
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            df.to_excel(file_path, sheet_name=sheet_name, index=index)
            logger.info(f"DataFrame salvo em {file_path}")
            
        except Exception as e:
            logger.error(f"Erro ao salvar DataFrame em {file_path}: {str(e)}")
            raise
    
    def create_formatted_excel(self, data: Dict[str, pd.DataFrame], 
                             file_path: str,
                             template_path: Optional[str] = None) -> None:
        """
        Criar arquivo Excel formatado
        
        Args:
            data: Dicionário com nome da aba e DataFrame
            file_path: Caminho do arquivo de saída
            template_path: Caminho do template (opcional)
        """
        try:
            # Criar workbook
            if template_path and os.path.exists(template_path):
                self.workbook = openpyxl.load_workbook(template_path)
            else:
                self.workbook = openpyxl.Workbook()
                # Remover aba padrão
                if "Sheet" in self.workbook.sheetnames:
                    self.workbook.remove(self.workbook["Sheet"])
            
            # Adicionar dados
            for sheet_name, df in data.items():
                if sheet_name in self.workbook.sheetnames:
                    worksheet = self.workbook[sheet_name]
                else:
                    worksheet = self.workbook.create_sheet(sheet_name)
                
                # Limpar dados existentes
                worksheet.delete_rows(1, worksheet.max_row)
                
                # Adicionar dados
                for r in dataframe_to_rows(df, index=False, header=True):
                    worksheet.append(r)
                
                # Aplicar formatação
                self._apply_formatting(worksheet, df)
            
            # Salvar arquivo
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            self.workbook.save(file_path)
            logger.info(f"Arquivo Excel formatado salvo em {file_path}")
            
        except Exception as e:
            logger.error(f"Erro ao criar arquivo Excel formatado: {str(e)}")
            raise
    
    def _apply_formatting(self, worksheet, df: pd.DataFrame) -> None:
        """
        Aplicar formatação na planilha
        
        Args:
            worksheet: Planilha do openpyxl
            df: DataFrame com dados
        """
        try:
            # Definir estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Formatar cabeçalho
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Formatar dados
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Ajustar largura das colunas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.warning(f"Erro ao aplicar formatação: {str(e)}")
    
    def validate_excel_structure(self, file_path: str, 
                               expected_sheets: Optional[List[str]] = None,
                               required_columns: Optional[Dict[str, List[str]]] = None) -> Dict[str, Any]:
        """
        Validar estrutura do arquivo Excel
        
        Args:
            file_path: Caminho do arquivo
            expected_sheets: Lista de abas esperadas
            required_columns: Dicionário com aba e colunas obrigatórias
            
        Returns:
            Dicionário com resultados da validação
        """
        try:
            validation_result = {
                "file_path": file_path,
                "valid": True,
                "errors": [],
                "warnings": [],
                "sheets_found": [],
                "sheets_missing": [],
                "columns_missing": {}
            }
            
            # Verificar se arquivo existe
            if not os.path.exists(file_path):
                validation_result["valid"] = False
                validation_result["errors"].append(f"Arquivo não encontrado: {file_path}")
                return validation_result
            
            # Ler estrutura do arquivo
            excel_file = pd.ExcelFile(file_path)
            found_sheets = excel_file.sheet_names
            validation_result["sheets_found"] = found_sheets
            
            # Validar abas esperadas
            if expected_sheets:
                for sheet in expected_sheets:
                    if sheet not in found_sheets:
                        validation_result["sheets_missing"].append(sheet)
                        validation_result["warnings"].append(f"Aba '{sheet}' não encontrada")
            
            # Validar colunas obrigatórias
            if required_columns:
                for sheet_name, required_cols in required_columns.items():
                    if sheet_name in found_sheets:
                        df = pd.read_excel(file_path, sheet_name=sheet_name)
                        found_columns = df.columns.tolist()
                        missing_columns = [col for col in required_cols if col not in found_columns]
                        
                        if missing_columns:
                            validation_result["columns_missing"][sheet_name] = missing_columns
                            validation_result["warnings"].append(
                                f"Aba '{sheet_name}': colunas faltando: {missing_columns}"
                            )
            
            # Verificar se há erros críticos
            if validation_result["sheets_missing"] and expected_sheets:
                validation_result["valid"] = False
                validation_result["errors"].append("Abas obrigatórias não encontradas")
            
            return validation_result
            
        except Exception as e:
            logger.error(f"Erro ao validar estrutura do Excel: {str(e)}")
            return {
                "file_path": file_path,
                "valid": False,
                "errors": [f"Erro na validação: {str(e)}"],
                "warnings": [],
                "sheets_found": [],
                "sheets_missing": [],
                "columns_missing": {}
            }
    
    def get_excel_info(self, file_path: str) -> Dict[str, Any]:
        """
        Obter informações do arquivo Excel
        
        Args:
            file_path: Caminho do arquivo
            
        Returns:
            Dicionário com informações do arquivo
        """
        try:
            info = {
                "file_path": file_path,
                "file_size": os.path.getsize(file_path),
                "last_modified": datetime.fromtimestamp(os.path.getmtime(file_path)),
                "sheets": {},
                "total_rows": 0,
                "total_columns": 0
            }
            
            excel_file = pd.ExcelFile(file_path)
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                info["sheets"][sheet_name] = {
                    "rows": len(df),
                    "columns": len(df.columns),
                    "column_names": df.columns.tolist()
                }
                info["total_rows"] += len(df)
                info["total_columns"] = max(info["total_columns"], len(df.columns))
            
            return info
            
        except Exception as e:
            logger.error(f"Erro ao obter informações do Excel: {str(e)}")
            raise
